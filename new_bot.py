from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO  # Для отправки файла без сохранения на диск
import logging

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Словарь для хранения информации от пользователей
user_data = {}

# Список районов Могилевской области
regions = [
    "Бобруйский район", "Быховский район", "Глусский район", "Горецкий район",
    "Дрибинский район", "Кировский район", "Климовичский район", "Кличевский район",
    "Костюковичский район", "Краснопольский район", "Кричевский район", "Круглянский район",
    "Могилевский район", "Мстиславский район", "Осиповичский район", "Славгородский район",
    "Хотимский район", "Чаусский район", "Чериковский район", "Шкловский район"
]

# ID администратора (замените на ваш Telegram ID)
ADMIN_ID = 513333176  # Замените на ваш реальный Telegram ID

# Функция для команды /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        keyboard = [[InlineKeyboardButton(region, callback_data=region)] for region in regions]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await update.message.reply_text('Выберите ваш район:', reply_markup=reply_markup)
    except Exception as e:
        logger.error(f"Ошибка в команде /start: {e}")

# Обработчик для выбора района
async def region_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        query = update.callback_query
        await query.answer()

        # Получаем выбранный район
        region = query.data
        user_id = query.from_user.id

        # Сохраняем район в словарь
        if user_id not in user_data:
            user_data[user_id] = {}

        user_data[user_id]['region'] = region

        # Кнопки для подтверждения района
        keyboard = [
            [InlineKeyboardButton("Да", callback_data="confirm_yes"),
             InlineKeyboardButton("Нет", callback_data="confirm_no")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        # Запрашиваем подтверждение у пользователя
        await query.edit_message_text(f"Вы выбрали {region}. Подтверждаете?", reply_markup=reply_markup)
    except Exception as e:
        logger.error(f"Ошибка в обработчике выбора района: {e}")

# Обработчик для подтверждения выбора региона
async def confirm_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        query = update.callback_query
        await query.answer()

        user_id = query.from_user.id

        # Если пользователь подтвердил выбор
        if query.data == "confirm_yes":
            await query.edit_message_text(f"Район {user_data[user_id]['region']} подтвержден.\n"
                                          f"Введите количество МП находящихся в производстве (в том числе приостановленных):")
            user_data[user_id]['step'] = 'mp_in_production'

        # Если пользователь отказался и хочет выбрать район заново
        elif query.data == "confirm_no":
            keyboard = [[InlineKeyboardButton(region, callback_data=region)] for region in regions]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.edit_message_text('Выберите ваш район еще раз:', reply_markup=reply_markup)
    except Exception as e:
        logger.error(f"Ошибка в обработчике подтверждения выбора региона: {e}")

# Обработчик для обработки числовых ответов
async def handle_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        user_id = update.message.from_user.id
        text = update.message.text

        # Проверка на то, ввел ли пользователь число
        if not text.isdigit():
            await update.message.reply_text('Пожалуйста, введите числовое значение.')
            return

        # Преобразуем текст в число
        number = int(text)

        # Проверяем, на каком шаге находится пользователь
        step = user_data[user_id].get('step')

        if step == 'mp_in_production':
            user_data[user_id]['mp_in_production'] = number
            user_data[user_id]['step'] = 'mp_suspended'
            await update.message.reply_text('Сколько из них приостановлено?')

        elif step == 'mp_suspended':
            user_data[user_id]['mp_suspended'] = number
            user_data[user_id]['step'] = 'mp_transferred'
            await update.message.reply_text('Сколько МП передано следственным подразделениям?')

        elif step == 'mp_transferred':
            user_data[user_id]['mp_transferred'] = number
            user_data[user_id]['step'] = None  # Завершаем сбор данных

            # Собранные данные
            mp_in_production = user_data[user_id].get('mp_in_production', 0)
            mp_suspended = user_data[user_id].get('mp_suspended', 0)
            mp_transferred = user_data[user_id].get('mp_transferred', 0)

            await update.message.reply_text(f"Спасибо! Вот собранная информация:\n"
                                            f"МП в производстве: {mp_in_production}\n"
                                            f"Из них приостановлено: {mp_suspended}\n"
                                            f"Передано следственным подразделениям: {mp_transferred}")
    except Exception as e:
        logger.error(f"Ошибка в обработчике ввода: {e}")

# Функция для команды /show_info
async def show_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        if user_data:
            info_message = "Собранная информация:\n"
            for user_id, data in user_data.items():
                region = data.get('region', 'Не выбран')
                mp_in_production = data.get('mp_in_production', 'Не указано')
                mp_suspended = data.get('mp_suspended', 'Не указано')
                mp_transferred = data.get('mp_transferred', 'Не указано')
                info_message += (f"Пользователь {user_id} - Район: {region}, "
                                 f"МП в производстве: {mp_in_production}, "
                                 f"Приостановлено: {mp_suspended}, "
                                 f"Передано следственным подразделениям: {mp_transferred}\n")
            await update.message.reply_text(info_message)
        else:
            await update.message.reply_text('Нет сохраненной информации.')
    except Exception as e:
        logger.error(f"Ошибка в команде /show_info: {e}")

# Функция для команды /export
async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        user_id = update.message.from_user.id

        # Проверяем, является ли пользователь администратором
        if user_id != ADMIN_ID:
            await update.message.reply_text('У вас нет прав для выполнения этой команды.')
            return

        if not user_data:
            await update.message.reply_text('Нет данных для экспорта.')
            return

        # Создаем DataFrame из user_data
        data = []
        for data_entry in user_data.values():
            if 'region' in data_entry:
                data.append({
                    'Район': data_entry['region'],
                    'МП в производстве': data_entry.get('mp_in_production', 0),
                    'Приостановлено': data_entry.get('mp_suspended', 0),
                    'Передано следственным подразделениям': data_entry.get('mp_transferred', 0)
                })

        df = pd.DataFrame(data)

        # Добавляем строку с общими суммами
        total_row = pd.DataFrame({
            'Район': ['Итого'],
            'МП в производстве': [df['МП в производстве'].sum()],
            'Приостановлено': [df['Приостановлено'].sum()],
            'Передано следственным подразделениям': [df['Передано следственным подразделениям'].sum()]
        })

        df = pd.concat([df, total_row], ignore_index=True)

        # Определяем районы, которые не предоставили отчет
        reported_regions = [entry['region'] for entry in user_data.values()]
        missing_regions = set(regions) - set(reported_regions)

        # Добавляем отсутствующие районы с нулевыми значениями
        missing_data = []
        for region in missing_regions:
            missing_data.append({
                'Район': region,
                'МП в производстве': 0,
                'Приостановлено': 0,
                'Передано следственным подразделениям': 0
            })

        if missing_data:
            missing_df = pd.DataFrame(missing_data)
            df = pd.concat([df, missing_df], ignore_index=True)

        # Сохраняем DataFrame в Excel-файл в памяти
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Report')
            # Удаляем вызов writer.save()

        # Загружаем книгу для стилизации
        excel_buffer.seek(0)  # Перемещаем указатель в начало файла
        wb = load_workbook(excel_buffer)
        ws = wb['Report']

        # Определяем красный цвет для подсветки
        red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')

        # Подсвечиваем строки, где 'МП в производстве' == 0 и Район не 'Итого'
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - len(missing_regions)):
            region_cell = row[0]  # Первый столбец: 'Район'
            mp_in_production_cell = row[1]  # Второй столбец: 'МП в производстве'

            if mp_in_production_cell.value == 0 and region_cell.value != 'Итого':
                for cell in row:
                    cell.fill = red_fill

        # Сохраняем изменения в буфер
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Отправляем Excel-файл администратору
        await update.message.reply_document(document=excel_buffer, filename='report.xlsx')
    except Exception as e:
        logger.error(f"Ошибка в команде /export: {e}")
        await update.message.reply_text('Произошла ошибка при экспорте данных. Попробуйте позже.')

# Функция main для запуска бота
def main():
    # Замените 'YOUR_BOT_TOKEN_HERE' на токен вашего бота, полученный от BotFather
    TOKEN = '7440834305:AAFy-1Gtx9-kRN0cn87dLyOpkJ8Rh3DKTG0'

    # Создаем приложение
    application = ApplicationBuilder().token(TOKEN).build()

    # Регистрируем обработчики команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("show_info", show_info))
    application.add_handler(CommandHandler("export", export_data))

    # Регистрируем обработчики CallbackQuery для кнопок
    application.add_handler(CallbackQueryHandler(region_choice, pattern='^(' + '|'.join(regions) + ')$'))
    application.add_handler(CallbackQueryHandler(confirm_choice, pattern='^confirm_(yes|no)$'))

    # Регистрируем обработчик сообщений для ввода чисел
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_input))

    # Запускаем бота
    try:
        logger.info("Бот запущен")
        application.run_polling()
    except Exception as e:
        logger.error(f"Ошибка при запуске бота: {e}")

if __name__ == '__main__':
    main()


